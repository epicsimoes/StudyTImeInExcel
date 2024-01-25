from pynput import keyboard
import time
from openpyxl import load_workbook

time_counter = 0.0
stopProgram = False
wb = load_workbook("test.xlsx")
ws = wb.active
localtime = time.strftime("%d/%m/%Y")


def check_for_end(key):
    global stopProgram
    if key == keyboard.Key.ctrl_r:  # if the right ctrl key is pressed then the loop will break
        try:
            stopProgram = True
            wb.save("test.xlsx")
        except PermissionError:
            # if the Excel file is open then the program will tell the user to close it
            print("Please close the excel file before running the program again")


def main():
    global time_counter, stopProgram, wb

    listener = keyboard.Listener(on_press=check_for_end, on_release=None)
    listener.start()

    while not stopProgram:
        time.sleep(1)
        time_counter += 1.0  # adds 1 to the time counter every second
        write_to_excel()


def write_to_excel():
    global time_counter, wb, ws
    # check if the Excel file is empty
    if count_cells() == 0:
        ws["B2"] = "Date"
        ws[f'B{count_cells() + 2}'] = localtime
        ws["C2"] = "Time(s)"
        ws[f"C{count_cells() + 1}"] = time_counter
    # if the date is the same as the last entry then it will add the time to the last entry
    if ws[f"B{count_cells() + 1}"].value == localtime:
        ws[f"C{count_cells() + 1}"] = int(ws[f"C{count_cells() + 1}"].value) + + 1


def count_cells():
    global ws
    count = 0
    for cell in ws['B']:
        if cell.value is not None:
            count += 1
    return count


if __name__ == '__main__':
    main()
