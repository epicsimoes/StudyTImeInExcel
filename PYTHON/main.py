from pynput import keyboard
import time
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, series

time_counter = 0.0
stopProgram = False
wb = load_workbook("test.xlsx")
ws = wb.active
localTime = time.strftime("%d/%m/%Y")
loadLastEntry = ""


def check_for_end(key):
    global stopProgram
    if key == keyboard.Key.ctrl_r:  # if the right ctrl key is pressed then the loop will break
        try:
            stopProgram = True
            create_chart()
            wb.save("test.xlsx")
        except PermissionError:
            # if the Excel file is open then the program will tell the user to close it
            print("Please close the excel file before running the program again")


def main():
    global time_counter, stopProgram, wb

    load()

    listener = keyboard.Listener(on_press=check_for_end, on_release=None)
    listener.start()

    while not stopProgram:
        time.sleep(1)
        time_counter += 1.0  # adds 1 to the time counter every second
        write_to_excel()


def write_to_excel():
    global time_counter, wb, ws, loadLastEntry
    # check if the B, C or the D column is empty
    # if it is then it will add the headers and the first entry
    if count_cells("B") == 0 or count_cells("C") == 0 or count_cells("D") == 0:
        add_headers()
        add_entry()
    # if the date is the same as the last entry then it will add the time to the last entry
    if check_for_last_entry():
        refactor_cells()
        load()
        ws[f'C{count_cells("B") + 1}'] = float(loadLastEntry) + time_counter
        ws[f'D{count_cells("D") + 1}'] = (float(loadLastEntry) + time_counter) / 3600
    else:  # if not then it will add a new entry
        add_entry()
        refactor_cells()


def count_cells(x):
    global ws
    count = 0
    for cell in ws[x]:
        if cell.value is not None:
            count += 1
    return count


def check_for_last_entry():
    # checks if the column has an entry for the current date in the B column
    global ws
    if ws[f'B{count_cells("B") + 1}'].value == localTime:
        return True
    return False


def load():
    global wb, ws, loadLastEntry
    loadLastEntry = str(ws[f'C{count_cells("C")+ 1}'].value)


def refactor_cells():
    # refactor the cells in the B column so that the cells don´t end in .0
    for cell in ws["C"]:
        if cell.value is not None and str(cell.value).endswith(".0"):
            cell.value = int((str(cell.value).replace(".0", "")))


def add_headers():
    global ws
    ws["B2"] = "Date"
    ws["C2"] = "Time(s)"
    ws["D2"] = "Time(h)"


def add_entry():
    global ws
    ws[f'B{count_cells("B") + 2}'] = localTime
    ws[f'C{count_cells("C") + 2}'] = time_counter
    ws[f'D{count_cells("D") + 2}'] = time_counter / 3600


def create_chart():
    global wb, ws
    chart = ws.charts[0]
    data = Reference(ws, min_col=4, min_row=2, max_row=count_cells("D") + 1)
    cat =  Reference(ws, min_col=2, min_row=3, max_row=count_cells("B") + 1)

if __name__ == '__main__':
    main()
